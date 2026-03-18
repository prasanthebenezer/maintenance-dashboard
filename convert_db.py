"""
convert_db.py — Migrate ARASCO Dashboard Excel data into SQLite

Usage:
    python convert_db.py                            # Default: ARASCO Dashboard Data.xlsx → dashboard.db
    python convert_db.py --xlsx path/to/file.xlsx   # Specify input Excel
    python convert_db.py --db path/to/dashboard.db  # Specify output database
    python convert_db.py --append                   # Append instead of replacing
"""

import sqlite3
import argparse
import json
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl

BASE_DIR = Path(__file__).parent
DEFAULT_XLSX = BASE_DIR / "ARASCO Dashboard Data.xlsx"
DEFAULT_DB = BASE_DIR / "dashboard.db"

# ─── Schema ──────────────────────────────────────────────

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS line_performance (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    date TEXT,
    site_id TEXT NOT NULL,
    line_id TEXT NOT NULL,
    running_time REAL DEFAULT 0,
    downtime_electrical REAL DEFAULT 0,
    downtime_mechanical REAL DEFAULT 0,
    downtime_utilities REAL DEFAULT 0
);
CREATE INDEX IF NOT EXISTS idx_lp_site ON line_performance(site_id);
CREATE INDEX IF NOT EXISTS idx_lp_date ON line_performance(date);

CREATE TABLE IF NOT EXISTS breakdown (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    date TEXT,
    site_id TEXT NOT NULL,
    line_id TEXT,
    equipment_id TEXT,
    breakdown_duration REAL DEFAULT 0,
    breakdown_reason TEXT,
    corrective_action TEXT,
    preventive_action TEXT,
    maintenance_issue TEXT DEFAULT 'Y',
    type TEXT
);
CREATE INDEX IF NOT EXISTS idx_bd_site ON breakdown(site_id);
CREATE INDEX IF NOT EXISTS idx_bd_date ON breakdown(date);

CREATE TABLE IF NOT EXISTS spare_parts_cost (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    date TEXT,
    site_id TEXT NOT NULL,
    budget_cost REAL DEFAULT 0,
    sp_cost REAL DEFAULT 0
);
CREATE INDEX IF NOT EXISTS idx_sp_site ON spare_parts_cost(site_id);

CREATE TABLE IF NOT EXISTS site_maintenance (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    date TEXT,
    site_id TEXT NOT NULL,
    pm_scheduled INTEGER DEFAULT 0,
    pm_completed INTEGER DEFAULT 0,
    corrective_tasks INTEGER DEFAULT 0
);
CREATE INDEX IF NOT EXISTS idx_sm_site ON site_maintenance(site_id);

CREATE TABLE IF NOT EXISTS rca (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    incident_date TEXT,
    site_id TEXT NOT NULL,
    incident_id TEXT,
    completed TEXT
);
CREATE INDEX IF NOT EXISTS idx_rca_site ON rca(site_id);
"""

# ─── Helpers ─────────────────────────────────────────────

def col_find(row_dict, *names):
    """Case-insensitive, whitespace-trimming column lookup (mirrors JS col())."""
    keys = list(row_dict.keys())
    for name in names:
        lower = name.lower()
        for k in keys:
            if lower in k.strip().lower():
                return row_dict[k]
    return None


def to_float(val, default=0.0):
    """Safely convert to float."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def to_int(val, default=0):
    """Safely convert to int."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def normalize_date(dt):
    """
    Fix SheetJS-style timezone drift: if time is past noon,
    round up to next day. Returns ISO date string (YYYY-MM-DD).
    """
    if dt is None:
        return None
    if isinstance(dt, datetime):
        if dt.hour >= 12:
            dt = dt + timedelta(days=1)
        return dt.strftime('%Y-%m-%d')
    if isinstance(dt, (int, float)):
        # Excel serial date number (days since Dec 30, 1899)
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=int(dt))
        return dt.strftime('%Y-%m-%d')
    # Try parsing string
    s = str(dt).strip()
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%dT%H:%M:%S'):
        try:
            parsed = datetime.strptime(s, fmt)
            return parsed.strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


MONTH_MAP = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}


def parse_mmm_yy(val):
    """
    Parse MMM/YY or similar month-year format to YYYY-MM-01.
    Also handles datetime objects (first-of-month dates from Excel).
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return normalize_date(val)
    if isinstance(val, (int, float)):
        return normalize_date(val)
    s = str(val).strip()
    # Try "MMM/YY" or "MMM-YY"
    for sep in ('/', '-', ' '):
        parts = s.split(sep)
        if len(parts) == 2:
            mmm = parts[0].strip().lower()[:3]
            yy = parts[1].strip()
            if mmm in MONTH_MAP:
                try:
                    year = int(yy)
                    if year < 100:
                        year += 2000
                    return f'{year}-{MONTH_MAP[mmm]:02d}-01'
                except ValueError:
                    continue
    # Fallback to general date parse
    return normalize_date(val)


def sheet_to_dicts(wb, sheet_name):
    """Convert a worksheet to list of dicts with header row as keys."""
    if sheet_name not in wb.sheetnames:
        print(f"  Warning: Sheet '{sheet_name}' not found in workbook")
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = val
        result.append(d)
    return result


# ─── Import Functions ────────────────────────────────────

def import_line_performance(conn, wb):
    rows = sheet_to_dicts(wb, 'Line_Performance')
    count = 0
    for r in rows:
        date = normalize_date(col_find(r, 'Date'))
        site_id = (str(col_find(r, 'Site_ID') or '')).strip()
        line_id = (str(col_find(r, 'Line_ID') or '')).strip()
        if not site_id:
            continue
        conn.execute(
            "INSERT INTO line_performance (date, site_id, line_id, running_time, downtime_electrical, downtime_mechanical, downtime_utilities) VALUES (?,?,?,?,?,?,?)",
            (date, site_id, line_id,
             to_float(col_find(r, 'Running_Time', 'Running Time')),
             to_float(col_find(r, 'Downtime_Electrical', 'Downtime Electrical')),
             to_float(col_find(r, 'Downtime_Mechanical', 'Downtime Mechanical')),
             to_float(col_find(r, 'Downtime_Utilities', 'Downtime Utilities')))
        )
        count += 1
    return count


def import_breakdown(conn, wb):
    rows = sheet_to_dicts(wb, 'Breakdown')
    count = 0
    for r in rows:
        date = normalize_date(col_find(r, 'Date'))
        site_id = (str(col_find(r, 'Site_ID') or '')).strip()
        if not site_id:
            continue
        conn.execute(
            "INSERT INTO breakdown (date, site_id, line_id, equipment_id, breakdown_duration, breakdown_reason, corrective_action, preventive_action, maintenance_issue, type) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (date, site_id,
             (str(col_find(r, 'Line_ID') or '')).strip(),
             str(col_find(r, 'Equipment_ID') or ''),
             to_float(col_find(r, 'Total_Breakdown_Minutes', 'Breakdown_Duration', 'Total_Breakdown', 'Total Breakdown')),
             str(col_find(r, 'Breakdown_Reason', 'Breakdown Reason') or ''),
             str(col_find(r, 'Corrective_Action', 'Corrective Action') or ''),
             str(col_find(r, 'Preventive_Action', 'Preventive Action') or ''),
             str(col_find(r, 'Maintenance_Issue', 'Maintenance Issue') or 'Y').strip(),
             str(col_find(r, 'Type') or ''))
        )
        count += 1
    return count


def import_spare_parts(conn, wb):
    rows = sheet_to_dicts(wb, 'Spare_Parts_Cost')
    count = 0
    for r in rows:
        date = parse_mmm_yy(col_find(r, 'MMM/YY', 'MMM', 'Date', 'Month'))
        site_id = (str(col_find(r, 'Site_ID') or '')).strip()
        if not site_id:
            continue
        # Use 'Total' first (includes SP + R&M + Building), fallback to SP_Cost
        sp_cost = to_float(col_find(r, 'Total', 'SP_Cost'))
        budget_cost = to_float(col_find(r, 'Budget_Cost', 'Budget'))
        conn.execute(
            "INSERT INTO spare_parts_cost (date, site_id, budget_cost, sp_cost) VALUES (?,?,?,?)",
            (date, site_id, budget_cost, sp_cost)
        )
        count += 1
    return count


def import_site_maintenance(conn, wb):
    rows = sheet_to_dicts(wb, 'Site_Maintenance')
    count = 0
    for r in rows:
        date = parse_mmm_yy(col_find(r, 'MMM/YY', 'MMM', 'Date', 'Month'))
        site_id = (str(col_find(r, 'Site_ID') or '')).strip()
        if not site_id:
            continue
        conn.execute(
            "INSERT INTO site_maintenance (date, site_id, pm_scheduled, pm_completed, corrective_tasks) VALUES (?,?,?,?,?)",
            (date, site_id,
             to_int(col_find(r, 'PM_Scheduled', 'Scheduled')),
             to_int(col_find(r, 'PM_Completed', 'Completed')),
             to_int(col_find(r, 'Corrective_tasks', 'Corrective')))
        )
        count += 1
    return count


def import_rca(conn, wb):
    rows = sheet_to_dicts(wb, 'RCA')
    count = 0
    for r in rows:
        date = normalize_date(col_find(r, 'Incident Date', 'Incident_Date', 'Date'))
        site_id = (str(col_find(r, 'Site_ID') or '')).strip()
        if not site_id:
            continue
        completed = str(col_find(r, 'Completed', 'Completed (Y/N)') or '').strip()
        incident_id = str(col_find(r, 'Incident ID', 'Incident_ID') or '').strip()
        conn.execute(
            "INSERT INTO rca (incident_date, site_id, incident_id, completed) VALUES (?,?,?,?)",
            (date, site_id, incident_id, completed)
        )
        count += 1
    return count


# ─── Main ────────────────────────────────────────────────

def create_tables(conn):
    """Create all tables (idempotent)."""
    conn.executescript(SCHEMA_SQL)


def create_tables_only(db_path):
    """Create tables in a new/existing database."""
    conn = sqlite3.connect(str(db_path))
    create_tables(conn)
    conn.commit()
    conn.close()
    print(f"Tables created in {db_path}")


def run_import(xlsx_path, db_path, append=False):
    """Import Excel data into SQLite."""
    xlsx_path = Path(xlsx_path)
    db_path = Path(db_path)

    if not xlsx_path.exists():
        print(f"Error: Excel file not found: {xlsx_path}")
        return False

    print(f"Opening {xlsx_path}...")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    print(f"  Sheets found: {wb.sheetnames}")

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA journal_mode=WAL")
    create_tables(conn)

    if not append:
        print("  Clearing existing data...")
        for table in ['line_performance', 'breakdown', 'spare_parts_cost', 'site_maintenance', 'rca']:
            conn.execute(f"DELETE FROM {table}")

    results = {}
    results['line_performance'] = import_line_performance(conn, wb)
    results['breakdown'] = import_breakdown(conn, wb)
    results['spare_parts_cost'] = import_spare_parts(conn, wb)
    results['site_maintenance'] = import_site_maintenance(conn, wb)
    results['rca'] = import_rca(conn, wb)

    conn.commit()
    conn.close()
    wb.close()

    print("\nImport complete:")
    for table, count in results.items():
        print(f"  {table}: {count} rows")
    print(f"\nDatabase: {db_path}")
    return True


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Convert ARASCO Excel to SQLite')
    parser.add_argument('--xlsx', default=str(DEFAULT_XLSX), help='Input Excel file')
    parser.add_argument('--db', default=str(DEFAULT_DB), help='Output SQLite database')
    parser.add_argument('--append', action='store_true', help='Append to existing data')
    args = parser.parse_args()

    success = run_import(args.xlsx, args.db, args.append)
    if not success:
        exit(1)
